import os
from typing import List, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from tasks import Task, ToDo, InProgress, Done, Status

FILE_PATH = "tasks.xlsx"

FILLS = {
    Status.TODO: PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE"),       # czerwony
    Status.IN_PROGRESS: PatternFill(fill_type="solid", start_color="FFEB9C", end_color="FFEB9C"),# żółty
    Status.DONE: PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE"),      # zielony
}

BOLD_FONT = Font(bold=True)


def ensure_file():
    """Sprawdza istnienie pliku; jeśli nie istnieje, tworzy nowy arkusz z nagłówkami."""
    if not os.path.exists(FILE_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Zadania"
        ws.append(["Tytuł", "Status"]);
        wb.save(FILE_PATH)


def read_tasks() -> List[Task]:
    ensure_file()
    wb = load_workbook(FILE_PATH)
    ws = wb[wb.sheetnames[0]]

    tasks: List[Task] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        title, status = row
        if status == Status.TODO.value:
            tasks.append(ToDo(title))
        elif status == Status.IN_PROGRESS.value:
            tasks.append(InProgress(title))
        elif status == Status.DONE.value:
            tasks.append(Done(title))
    return tasks


def write_tasks(tasks: List[Task]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Zadania"
    ws.append(["Tytuł", "Status"])
    for t in tasks:
        ws.append([t.title, t.status.value])
    # formatowanie
    for row in ws.iter_rows(min_row=2):
        cell_title, cell_status = row
        status = cell_status.value
        # dobierz Enum
        enum_s = Status(status)
        # kolor
        cell_title.fill = FILLS[enum_s]
        cell_status.fill = FILLS[enum_s]
        # pogrubienie dla "W trakcie"
        if enum_s == Status.IN_PROGRESS:
            cell_title.font = BOLD_FONT
            cell_status.font = BOLD_FONT
    wb.save(FILE_PATH)


def add_task(title: str):
    tasks = read_tasks()
    tasks.append(ToDo(title))
    write_tasks(tasks)
    print(f"Dodano zadanie: '{title}' (Do zrobienia)")


def start_task(title: str):
    tasks = read_tasks()
    if any(t.status == Status.IN_PROGRESS for t in tasks):
        print("Błąd: Już istnieje zadanie w trakcie. Zakończ je przed rozpoczęciem nowego.")
        return
    for t in tasks:
        if t.title == title and t.status == Status.TODO:
            t.status = Status.IN_PROGRESS
            write_tasks(tasks)
            print(f"Rozpoczęto zadanie: '{title}'")
            return
    print(f"Nie znaleziono zadania '{title}' w statusie Do zrobienia.")


def finish_task(title: str):
    tasks = read_tasks()
    for t in tasks:
        if t.title == title and t.status == Status.IN_PROGRESS:
            t.status = Status.DONE
            write_tasks(tasks)
            print(f"Zakończono zadanie: '{title}'")
            # Inteligentna propozycja: pierwsze "Do zrobienia"
            todo_tasks = [x for x in tasks if x.status == Status.TODO]
            if todo_tasks:
                print(f"Możesz teraz rozpocząć zadanie: '{todo_tasks[0].title}'")
            return
    print(f"Nie znaleziono zadania '{title}' w statusie W trakcie.")


def show_tasks():
    from colorama import Fore, Style, init
    init(autoreset=True)
    tasks = read_tasks()
    print("\nZadania:")
    for status in Status:
        print(f"\n== {status.value} ==")
        for t in tasks:
            if t.status == status:
                text = t.title
                if status == Status.TODO:
                    print(Fore.RED + text)
                elif status == Status.IN_PROGRESS:
                    print(Style.BRIGHT + Fore.YELLOW + text)
                else:
                    print(Fore.GREEN + text)
    print()